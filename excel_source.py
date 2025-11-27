from __future__ import annotations

from pathlib import Path
from typing import Iterator, Optional, Tuple

from openpyxl import load_workbook


def _to_float_or_none(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    # Допускаем дробные через запятую
    s = s.replace(' ', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def read_excel_queries(path: str | Path) -> Iterator[Tuple[str, Optional[float]]]:
    """
    Читает Excel: колонка A — запросы/артикулы через запятую,
    колонка B — MaxPrice (USD), может быть пустым.

    Возвращает пары (query, max_price). Если max_price пуст — None.
    """
    p = Path(path)
    wb = load_workbook(p, data_only=True, read_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        a = row[0] if len(row) > 0 else None
        b = row[1] if len(row) > 1 else None
        if a is None:
            continue
        max_price = _to_float_or_none(b)
        for raw in str(a).split(','):
            q = raw.strip()
            if not q:
                continue
            yield q, max_price
